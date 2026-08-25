# -*- coding: utf-8 -*-
"""Gera o dashboard Pendencias de Sincronismo e o snapshot associado."""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import bindparam, create_engine, text
from sqlalchemy.engine import URL

from env_utils import load_env_file

WORKSPACE = Path(__file__).resolve().parent
SNAPSHOT_DIR = WORKSPACE / "snapshot_pendencias_sincronismo"
SNAPSHOT_JSON = SNAPSHOT_DIR / "pendencias_sincronismo.json"
TEMPLATE = WORKSPACE / "pendencias_sincronismo_template.html"
OUT_HTML = WORKSPACE / "PENDENCIAS_SINCRONISMO.html"
OUT_CSV = WORKSPACE / "PENDENCIAS_SINCRONISMO.csv"
OUT_XLSX = WORKSPACE / "PENDENCIAS_SINCRONISMO.xlsx"
OUT_MANIFEST = WORKSPACE / "MANIFESTO_SNAPSHOT_PENDENCIAS_SINCRONISMO.json"
SEED_HTML = Path(r"C:\Users\Administrador\Downloads\SEM_SYNC_POR_UF.html")
DESKTOP_DIR = Path(r"C:\Users\Administrador\Desktop\LISTA SEM SINCRONIZAÇÃO")
REVERSA_HTML = WORKSPACE / "REVERSA_DATALOGGERS.html"
SYNC_TOLERANCE = pd.Timedelta(minutes=15)

EXPORT_COLS = [
    "Pedido",
    "Logger",
    "UF",
    "Coleta",
    "Entrega",
    "Último Sync",
    "Dias sem sync",
    "Em GRU?",
    "Localização",
    "Situação atual",
    "Responsável atual",
    "Tipo",
    "Atualização dtbPortal",
    "Ação sugerida",
    "Última ação portal",
    "Último histórico portal",
    "Movimentos portal",
    "LPN",
    "Chegada cliente",
    "CTE",
    "Observação portal",
    "Tag no dtbPortal",
]


def now_brt() -> datetime:
    return pd.Timestamp.now(tz="America/Sao_Paulo").tz_localize(None).to_pydatetime()


def fmt_stamp(value: datetime | pd.Timestamp | None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    ts = pd.Timestamp(value)
    if pd.isna(ts):
        return ""
    return ts.strftime("%d/%m/%Y %H:%M")


def parse_br(value: object) -> pd.Timestamp | pd.NaT:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value if value.tzinfo is None else value.tz_convert("America/Sao_Paulo").tz_localize(None)
    raw = str(value).strip()
    if not raw or raw.lower() in {"nan", "nat", "none"}:
        return pd.NaT
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%y %H:%M", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return pd.Timestamp(datetime.strptime(raw[:19], fmt))
        except ValueError:
            continue
    return pd.to_datetime(raw, dayfirst=True, errors="coerce")


def load_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def norm_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text_value = unicodedata.normalize("NFD", str(value))
    text_value = "".join(ch for ch in text_value if unicodedata.category(ch) != "Mn")
    return " ".join(text_value.strip().upper().split())


def norm_tag(value: object) -> str:
    tag = re.sub(r"\s+", "", norm_text(value)).replace("-", "")
    match = re.fullmatch(r"S(\d+)", tag)
    return "S" + match.group(1).zfill(4) if match else tag


def acao_sugerida(localizacao: str, situacao: str) -> str:
    loc = norm_text(localizacao)
    sit = norm_text(situacao)
    if loc == "EM GRU":
        return "SINCRONIZAR EM GRU"
    if loc.startswith("RETORNANDO"):
        return "ACOMPANHAR RECEBIMENTO EM GRU"
    if loc.startswith("NAO LOCALIZADO") or loc.startswith("NÃO LOCALIZADO"):
        return "VALIDAR CADASTRO/POSIÇÃO NO DTBPORTAL"
    if sit == "AGENTE":
        return "ACIONAR AGENTE/RESPONSÁVEL"
    if "MANUTEN" in sit:
        return "VALIDAR COM MANUTENÇÃO"
    if "CAMARA" in sit or "CÂMARA" in sit:
        return "VALIDAR NA CÂMARA FRIA"
    return "VALIDAR COM RESPONSÁVEL ATUAL"


def extract_json_array(text: str, prefix: str) -> list:
    match = re.search(re.escape(prefix) + r"\s*=\s*(\[.*?\]);", text, re.S)
    if not match:
        return []
    return json.loads(match.group(1))


def load_seed_records() -> tuple[list[dict], str]:
    if SNAPSHOT_JSON.exists():
        payload = json.loads(SNAPSHOT_JSON.read_text(encoding="utf-8"))
        records = payload.get("dados") or []
        snapshot_em = str(payload.get("snapshot_em") or "")
        if records:
            return records, snapshot_em
    if SEED_HTML.exists():
        html = SEED_HTML.read_text(encoding="utf-8")
        records = extract_json_array(html, "const DADOS")
        stamp = "2026-08-25T11:52:00"
        match = re.search(r"Atualizado em ([0-9/]+ [0-9:]+)", html)
        if match:
            parsed = parse_br(match.group(1))
            if pd.notna(parsed):
                stamp = parsed.strftime("%Y-%m-%dT%H:%M:%S")
        return records, stamp
    raise FileNotFoundError("Snapshot de pendencias de sincronismo nao encontrado.")


def records_from_audit_xlsx(path: Path) -> list[dict]:
    source = pd.read_excel(path, sheet_name="SEM_SYNC_ATUAL")
    env = load_env(WORKSPACE / ".env.vtc_stage")
    uf_map: dict[str, str] = {}
    pedidos = sorted({str(p).strip() for p in source["Pedido"].dropna()})
    if env.get("VTC_STAGE_HOST") and pedidos:
        url = URL.create(
            "postgresql+psycopg2",
            username=env["VTC_STAGE_USER"],
            password=env["VTC_STAGE_PASSWORD"],
            host=env["VTC_STAGE_HOST"],
            port=int(env.get("VTC_STAGE_PORT") or 5432),
            database=env["VTC_STAGE_NAME"],
        )
        query = text(
            """
            SELECT TRIM(nr_pedido::text) AS pedido, MAX(NULLIF(TRIM(cd_uf), '')) AS uf
            FROM vtc_stage.documentos
            WHERE TRIM(nr_pedido::text) IN :peds
            GROUP BY TRIM(nr_pedido::text)
            """
        ).bindparams(bindparam("peds", expanding=True))
        engine = create_engine(url, pool_pre_ping=True)
        with engine.connect() as connection:
            for row in connection.execute(query, {"peds": pedidos}):
                if row[1]:
                    uf_map[str(row[0])] = str(row[1]).strip().upper()

    def fmt(value: object) -> str:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%d/%m/%y %H:%M")

    records = []
    for _, row in source.iterrows():
        localizacao = str(row.get("Classificação GRU") or "NÃO INFORMADO")
        situacao = str(row.get("Situação oficial dtbPortal") or "NÃO LOCALIZADO")
        pedido = str(row.get("Pedido") or "").strip()
        records.append(
            {
                "Pedido": pedido,
                "Logger": str(row.get("Logger") or "").strip(),
                "UF": uf_map.get(pedido, ""),
                "Coleta": fmt(row.get("Coleta efetiva")),
                "Entrega": fmt(row.get("Entrega efetiva")),
                "Último Sync": fmt(row.get("Último Sync (BRT)")),
                "Dias sem sync": int(row["Dias desde entrega"]) if pd.notna(row.get("Dias desde entrega")) else 0,
                "Em GRU?": str(row.get("Em GRU") or "NÃO"),
                "Localização": localizacao,
                "Situação atual": situacao,
                "Responsável atual": str(row.get("Responsável atual dtbPortal") or ""),
                "Tipo": str(row.get("Tipo datalogger") or ""),
                "Atualização dtbPortal": fmt(row.get("Atualização dtbPortal")),
                "Ação sugerida": acao_sugerida(localizacao, situacao),
                "Última ação portal": str(row.get("Última ação portal") or ""),
                "Último histórico portal": fmt(row.get("Último histórico portal")),
                "Movimentos portal": int(row["Movimentos no portal"]) if pd.notna(row.get("Movimentos no portal")) else 0,
                "LPN": row.get("LPN") if pd.notna(row.get("LPN")) else "",
                "Chegada cliente": fmt(row.get("Chegada ao cliente")),
                "CTE": str(row.get("CTE") or ""),
                "Observação portal": str(row.get("Observação última ação") or ""),
                "Tag no dtbPortal": str(row.get("Tag no dtbPortal") or ""),
            }
        )
    return records


def overlay_reversa(records: list[dict], pagina_em: datetime) -> tuple[list[dict], bool]:
    if not REVERSA_HTML.exists():
        return records, False
    html = REVERSA_HTML.read_text(encoding="utf-8", errors="replace")
    match = re.search(r"const ALL_ROWS=(\[.*?\]);", html)
    if not match:
        return records, False
    rows = json.loads(match.group(1))
    by_key: dict[tuple[str, str], list] = {}
    for row in rows:
        if len(row) < 12:
            continue
        key = (str(row[0]).strip(), norm_tag(row[2]))
        by_key[key] = row

    kept = []
    used = False
    for rec in records:
        key = (str(rec.get("Pedido") or "").strip(), norm_tag(rec.get("Logger")))
        row = by_key.get(key)
        if not row:
            kept.append(rec)
            continue
        used = True
        status = str(row[11] or "")
        if status == "Sincronizado":
            continue
        rec = dict(rec)
        if row[5]:
            rec["Entrega"] = row[5]
        if row[9]:
            rec["Último Sync"] = row[9]
        if row[13]:
            rec["UF"] = rec.get("UF") or row[13]
        elif row[12]:
            rec["UF"] = rec.get("UF") or row[12]
        if row[1]:
            rec["LPN"] = rec.get("LPN") or row[1]
        entrega = parse_br(rec.get("Entrega"))
        if pd.notna(entrega):
            rec["Dias sem sync"] = int((pd.Timestamp(pagina_em) - entrega).total_seconds() // 86400)
        kept.append(rec)
    return kept, used


def refresh_dtbportal(records: list[dict]) -> bool:
    load_env_file(WORKSPACE / ".env")
    tags = sorted({norm_tag(r.get("Logger")) for r in records if r.get("Logger")})
    if not tags:
        return False
    env = load_env(WORKSPACE / ".env")
    if not env.get("AURA_POSTGRES_HOST"):
        return False
    tag_norm = "REPLACE(REPLACE(UPPER(TRIM(vwt.ds_tag)), '-', ''), ' ', '')"
    query = text(
        f"""
        SELECT DISTINCT ON ({tag_norm})
               {tag_norm} AS logger,
               UPPER(TRIM(vwt.ds_tag)) AS tag_portal,
               vwt.ds_tipodatalogger,
               vwt.ds_destino,
               vwt.ds_finalidade,
               vwt.ds_responsavel,
               vwt.ds_statusrecebimento,
               vwt.dt_atualizacao
        FROM vwTabelaMovDataloggers vwt
        WHERE {tag_norm} IN :tags
        ORDER BY {tag_norm}, vwt.dt_atualizacao DESC NULLS LAST
        """
    ).bindparams(bindparam("tags", expanding=True))
    url = URL.create(
        "postgresql+psycopg2",
        username=env["AURA_POSTGRES_USER"],
        password=env["AURA_POSTGRES_PASSWORD"],
        host=env["AURA_POSTGRES_HOST"],
        port=int(env.get("AURA_POSTGRES_PORT") or 5432),
        database=env["AURA_POSTGRES_NAME"],
    )
    engine = create_engine(url, pool_pre_ping=True)
    with engine.connect() as connection:
        frame = pd.read_sql(query, connection, params={"tags": tags})
    if frame.empty:
        return False
    pos = {norm_tag(row.logger): row for row in frame.itertuples(index=False)}
    for rec in records:
        row = pos.get(norm_tag(rec.get("Logger")))
        if row is None:
            continue
        dest = norm_text(getattr(row, "ds_destino", ""))
        if "ESTOQUE" in dest and "GRU" in dest:
            rec["Localização"] = "EM GRU"
            rec["Em GRU?"] = "SIM"
        elif "RETORNANDO" in dest and "GRU" in dest:
            rec["Localização"] = "RETORNANDO PARA GRU (NÃO RECEBIDO)"
            rec["Em GRU?"] = "NÃO"
        else:
            rec["Localização"] = rec.get("Localização") or "FORA DE GRU"
            rec["Em GRU?"] = "NÃO"
        rec["Responsável atual"] = str(getattr(row, "ds_responsavel", "") or rec.get("Responsável atual") or "")
        rec["Tipo"] = rec.get("Tipo") or str(getattr(row, "ds_tipodatalogger", "") or "")
        rec["Tag no dtbPortal"] = str(getattr(row, "tag_portal", "") or rec.get("Tag no dtbPortal") or "")
        atual = pd.to_datetime(getattr(row, "dt_atualizacao", None), errors="coerce")
        if pd.notna(atual):
            rec["Atualização dtbPortal"] = atual.strftime("%d/%m/%y %H:%M")
        rec["Ação sugerida"] = acao_sugerida(str(rec.get("Localização") or ""), str(rec.get("Situação atual") or ""))
    return True


def build_resumo(records: list[dict]) -> list[dict]:
    grupos: dict[str, dict] = {}
    for rec in records:
        uf = rec.get("UF") or "SEM UF"
        item = grupos.setdefault(
            uf,
            {"UF": uf, "Pendências": 0, "Em GRU": 0, "Fora de GRU": 0, "Retornando p/ GRU": 0, "dias": []},
        )
        item["Pendências"] += 1
        dias = int(rec.get("Dias sem sync") or 0)
        item["dias"].append(dias)
        if rec.get("Em GRU?") == "SIM":
            item["Em GRU"] += 1
        elif str(rec.get("Localização") or "").startswith("RETORNANDO"):
            item["Retornando p/ GRU"] += 1
        else:
            item["Fora de GRU"] += 1
    out = []
    for item in grupos.values():
        dias = item.pop("dias")
        out.append(
            {
                **item,
                "Dias sem sync (máx)": max(dias) if dias else 0,
                "Dias sem sync (média)": round(sum(dias) / len(dias), 1) if dias else 0,
            }
        )
    return sorted(out, key=lambda row: (-int(row["Pendências"]), str(row["UF"])))


def write_excel(records: list[dict], resumo: list[dict], pagina_em: datetime, snapshot_em: str) -> None:
    cabecalho = pd.DataFrame(
        [
            ("Página atualizada em", fmt_stamp(pagina_em)),
            ("Snapshot em", snapshot_em),
            ("Total de pendências", len(records)),
            ("UFs com pendência", len({r.get("UF") for r in records if r.get("UF")})),
            ("Fontes visíveis", "dtbPortal · portal VTC"),
        ],
        columns=["Indicador", "Valor"],
    )
    detalhe = pd.DataFrame(records)
    for col in EXPORT_COLS:
        if col not in detalhe.columns:
            detalhe[col] = ""
    detalhe = detalhe[EXPORT_COLS]
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        cabecalho.to_excel(writer, sheet_name="RESUMO", index=False)
        pd.DataFrame(resumo).to_excel(writer, sheet_name="RESUMO_UF", index=False)
        detalhe.to_excel(writer, sheet_name="PENDENCIAS", index=False)
    workbook = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, cells in enumerate(sheet.columns, 1):
            values = [str(cell.value or "") for cell in list(cells)[:400]]
            width = min(max(max((len(value) for value in values), default=0) + 2, 10), 42)
            sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(OUT_XLSX)


def write_html(records: list[dict], pagina_em: datetime, snapshot_em: str) -> None:
    template = TEMPLATE.read_text(encoding="utf-8")
    pagina_txt = fmt_stamp(pagina_em)
    snap_txt = snapshot_em
    if re.match(r"\d{4}-\d{2}-\d{2}", snapshot_em):
        parsed = parse_br(snapshot_em.replace("T", " "))
        if pd.notna(parsed):
            snap_txt = fmt_stamp(parsed)
    meta = {
        "pagina_em": pagina_txt,
        "snapshot_em": snap_txt,
        "fontes": ["dtbPortal", "portal VTC"],
        "linhas": len(records),
    }
    html = (
        template.replace("__PAGINA_EM__", pagina_txt)
        .replace("__SNAPSHOT_EM__", snap_txt)
        .replace("__DADOS_JSON__", json.dumps(records, ensure_ascii=False, default=str))
        .replace("__META_JSON__", json.dumps(meta, ensure_ascii=False))
    )
    OUT_HTML.write_text(html, encoding="utf-8")


def main() -> None:
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    pagina_em = now_brt()
    live_bits: list[str] = []
    records, snapshot_em = load_seed_records()
    origem = "snapshot local"

    audit_files = []
    if DESKTOP_DIR.exists():
        audit_files = [
            path
            for path in DESKTOP_DIR.glob("SEM_SYNC_ATUALIZADO_*.xlsx")
            if not path.name.startswith("~$")
        ]
    if audit_files:
        latest = max(audit_files, key=lambda path: path.stat().st_mtime)
        try:
            records = records_from_audit_xlsx(latest)
            snapshot_em = pagina_em.strftime("%Y-%m-%dT%H:%M:%S")
            origem = latest.name
            live_bits.append("planilha auditada")
        except Exception as exc:
            print(f"AVISO: nao foi possivel ler {latest.name}: {exc}")

    try:
        records, used = overlay_reversa(records, pagina_em)
        if used:
            live_bits.append("reversa")
            snapshot_em = pagina_em.strftime("%Y-%m-%dT%H:%M:%S")
    except Exception as exc:
        print(f"AVISO: overlay da reversa nao aplicado: {exc}")

    try:
        if refresh_dtbportal(records):
            live_bits.append("dtbPortal")
            snapshot_em = pagina_em.strftime("%Y-%m-%dT%H:%M:%S")
    except Exception as exc:
        print(f"AVISO: dtbPortal nao atualizado: {exc}")

    for rec in records:
        rec["UF"] = rec.get("UF") or "SEM UF"
        rec["Dias sem sync"] = int(rec.get("Dias sem sync") or 0)
        for key, value in list(rec.items()):
            if value is None or (isinstance(value, float) and pd.isna(value)) or str(value).lower() == "nan":
                rec[key] = ""
        if rec.get("Ação sugerida"):
            rec["Ação sugerida"] = re.sub(r"\bESL\b", "portal VTC", str(rec["Ação sugerida"]), flags=re.I)

    resumo = build_resumo(records)
    detalhe = pd.DataFrame(records)
    for col in EXPORT_COLS:
        if col not in detalhe.columns:
            detalhe[col] = ""
    detalhe[EXPORT_COLS].to_csv(OUT_CSV, index=False, encoding="utf-8-sig", sep=";")
    write_excel(records, resumo, pagina_em, snapshot_em if "T" not in snapshot_em else fmt_stamp(parse_br(snapshot_em.replace("T", " "))))
    write_html(records, pagina_em, snapshot_em)

    payload = {
        "origem": origem,
        "snapshot_em": snapshot_em,
        "pagina_em": pagina_em.strftime("%Y-%m-%dT%H:%M:%S"),
        "linhas": len(records),
        "dados": records,
        "resumo": resumo,
    }
    SNAPSHOT_JSON.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    status = "VALIDADO_COM_FONTES_FRESCAS" if live_bits else "SNAPSHOT_REUTILIZADO"
    manifesto = {
        "gerado_em": pagina_em.strftime("%Y-%m-%dT%H:%M:%S"),
        "snapshot_em": snapshot_em,
        "status": status,
        "linhas": len(records),
        "ufs": len({r.get("UF") for r in records if r.get("UF") and r.get("UF") != "SEM UF"}),
        "fontes_visiveis": ["dtbPortal", "portal VTC"],
        "fontes_aplicadas": live_bits,
        "origem": origem,
        "arquivos": [
            OUT_HTML.name,
            OUT_CSV.name,
            OUT_XLSX.name,
        ],
    }
    OUT_MANIFEST.write_text(json.dumps(manifesto, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"STATUS {status}")
    print(f"HTML: {OUT_HTML}")
    print(f"CSV: {OUT_CSV}")
    print(f"XLSX: {OUT_XLSX}")
    print(f"Pendencias: {len(records)} | UFs: {manifesto['ufs']} | fontes: {', '.join(live_bits) or 'snapshot'}")


if __name__ == "__main__":
    main()
